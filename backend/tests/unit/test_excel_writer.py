"""Smoke + shape tests del template writer y exporter (post sections-widgets-redesign).

No cubren el parser. El foco es que el archivo que ve el operador
tenga las 11 hojas en orden, headers correctos, dropdowns en los enums
y que la hoja Instrucciones cubra A/B/C/D.
"""
from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from apps.reports.importers import schema as s
from apps.reports.importers.excel_exporter import export
from apps.reports.importers.excel_writer import build_template
from apps.reports.models import (
    ChartDataPoint,
    ChartWidget,
    ImageWidget,
    KpiGridWidget,
    KpiTile,
    Report,
    Section,
    TableRow,
    TableWidget,
    TextImageWidget,
    TextWidget,
    TopContentItem,
    TopContentsWidget,
    TopCreatorItem,
    TopCreatorsWidget,
)
from apps.reports.tests.factories import make_stage


# ---------------------------------------------------------------------------
# Template writer — no DB
# ---------------------------------------------------------------------------
def test_template_has_11_sheets_in_order():
    buf = build_template()
    wb = load_workbook(buf)
    assert wb.sheetnames == s.SHEETS_IN_ORDER


def test_template_instrucciones_covers_llm_and_human_sections():
    wb = load_workbook(build_template())
    ws = wb[s.SHEET_INSTRUCCIONES]
    text = "\n".join(
        str(ws.cell(row=r, column=1).value or "")
        for r in range(1, ws.max_row + 1)
    )
    assert "A. Cómo llenar el Excel" in text
    assert "B. Cómo armar el ZIP" in text
    assert "C. Regenerar el template" in text
    assert "D. Para LLMs" in text
    # Widget types mentioned
    assert "TextWidget" in text
    assert "TableWidget" in text
    # Image extension contract
    assert ".jpg" in text and ".png" in text


def test_template_reporte_sheet_has_kv_rows_without_layout():
    wb = load_workbook(build_template())
    ws = wb[s.SHEET_REPORTE]

    col_a = [str(ws.cell(row=r, column=1).value or "") for r in range(1, ws.max_row + 1)]
    assert any("tipo*" in v for v in col_a), "tipo debería estar marcado obligatorio"
    assert any("fecha_inicio*" in v for v in col_a)
    # No Layout table in new schema
    assert not any("# Layout" in v for v in col_a), "Reporte ya no tiene tabla Layout"


def test_template_sections_sheet_has_exact_headers():
    wb = load_workbook(build_template())
    ws = wb[s.SHEET_SECTIONS]
    actual = [ws.cell(row=1, column=c).value for c in range(1, len(s.SECTIONS_HEADERS) + 1)]
    assert actual == s.SECTIONS_HEADERS


def test_template_widget_sheets_have_exact_headers():
    wb = load_workbook(build_template())
    for sheet_name, expected_headers in s.SHEET_HEADERS.items():
        ws = wb[sheet_name]
        actual = [ws.cell(row=1, column=c).value for c in range(1, len(expected_headers) + 1)]
        assert actual == expected_headers, f"{sheet_name}: headers mismatch"


def test_template_has_dropdown_on_tipo():
    wb = load_workbook(build_template())
    ws = wb[s.SHEET_REPORTE]
    assert ws.data_validations.count > 0, "Reporte debería tener al menos un dropdown"


def test_template_dropdowns_on_widget_sheets_match_schema():
    wb = load_workbook(build_template())
    for (sheet_name, _header), _choices in s.DROPDOWNS.items():
        if sheet_name == s.SHEET_REPORTE:
            continue  # Reporte es key-value, validado arriba
        ws = wb[sheet_name]
        assert ws.data_validations.count > 0, (
            f"{sheet_name} debería tener al menos un dropdown"
        )


# ---------------------------------------------------------------------------
# Exporter — necesita DB
# ---------------------------------------------------------------------------
@pytest.fixture
def tiny_report(db):
    """Report con 1 Section y 1 widget de cada tipo (sin imágenes reales)."""
    stage = make_stage()
    report = Report.objects.create(
        stage=stage,
        kind=Report.Kind.MENSUAL,
        period_start=date(2026, 4, 1),
        period_end=date(2026, 4, 30),
        title="Reporte tiny",
        intro_text="Intro de prueba",
        conclusions_text="Conclusiones de prueba",
        status=Report.Status.DRAFT,
        published_at=datetime(2026, 5, 2, 12, 0, tzinfo=timezone.utc),
    )
    section = Section.objects.create(
        report=report, order=1, title="Análisis", layout="stack",
    )

    TextWidget.objects.create(section=section, order=1, title="Intro", body="Body text")
    TextImageWidget.objects.create(
        section=section, order=2, title="Hero",
        body="Body", image_position="left", columns=1,
    )

    kpi = KpiGridWidget.objects.create(section=section, order=3, title="KPIs")
    KpiTile.objects.create(
        widget=kpi, order=1, label="Reach",
        value=Decimal("1000"), period_comparison=Decimal("5.0"),
    )
    KpiTile.objects.create(
        widget=kpi, order=2, label="Engagement",
        value=Decimal("5.5"), period_comparison=None,
    )

    tbl = TableWidget.objects.create(section=section, order=4, title="Mes a mes")
    TableRow.objects.create(
        widget=tbl, order=1, is_header=True,
        cells=["Métrica", "Valor", "Δ"],
    )
    TableRow.objects.create(
        widget=tbl, order=2,
        cells=["reach", "500", "+5%"],
    )

    tc = TopContentsWidget.objects.create(
        section=section, order=5, title="Top posts",
        network="INSTAGRAM", period_label="abril",
    )
    TopContentItem.objects.create(
        widget=tc, order=1, caption="Post 1",
        source_type="ORGANIC", views=100, likes=10,
    )

    tcr = TopCreatorsWidget.objects.create(
        section=section, order=6, title="Top creators",
        network="INSTAGRAM", period_label="abril",
    )
    TopCreatorItem.objects.create(
        widget=tcr, order=1, handle="@test", views=200,
    )

    chart = ChartWidget.objects.create(
        section=section, order=7, title="Followers",
        network="INSTAGRAM", chart_type="bar",
    )
    ChartDataPoint.objects.create(
        widget=chart, order=1, label="Ene", value=Decimal("100"),
    )
    ChartDataPoint.objects.create(
        widget=chart, order=2, label="Feb", value=Decimal("120"),
    )

    return report


def test_exporter_returns_bytesio(tiny_report):
    buf = export(tiny_report)
    assert isinstance(buf, BytesIO)
    assert len(buf.getvalue()) > 0


def test_exporter_preserves_sheet_layout_from_template(tiny_report):
    """Shape = mismas hojas en el mismo orden que el template vacío."""
    wb = load_workbook(export(tiny_report))
    assert wb.sheetnames == s.SHEETS_IN_ORDER


def test_exporter_populates_reporte_kv(tiny_report):
    wb = load_workbook(export(tiny_report))
    ws = wb[s.SHEET_REPORTE]

    kv = {}
    for r in range(2, ws.max_row + 1):
        key = ws.cell(row=r, column=1).value
        if isinstance(key, str) and key and not key.startswith("#"):
            kv[key.rstrip("*")] = ws.cell(row=r, column=2).value

    assert kv["tipo"] == "Mensual"
    assert kv["fecha_inicio"] == "01/04/2026"
    assert kv["fecha_fin"] == "30/04/2026"
    assert kv["titulo"] == "Reporte tiny"
    assert kv["intro"] == "Intro de prueba"
    assert kv["conclusiones"] == "Conclusiones de prueba"


def test_exporter_populates_sections_sheet(tiny_report):
    wb = load_workbook(export(tiny_report))
    rows = _read_tabular(wb[s.SHEET_SECTIONS], s.SECTIONS_HEADERS)
    assert len(rows) == 1
    assert rows[0]["nombre"] == "section_1"
    assert rows[0]["title"] == "Análisis"
    assert rows[0]["layout"] == "stack"
    assert rows[0]["order"] == 1


def test_exporter_writes_kpi_tiles_denormalized(tiny_report):
    wb = load_workbook(export(tiny_report))
    rows = _read_tabular(wb[s.SHEET_KPIGRIDS], s.KPIGRIDS_HEADERS)
    assert len(rows) == 2
    assert all(r["section_nombre"] == "section_1" for r in rows)
    assert all(r["widget_title"] == "KPIs" for r in rows)
    assert rows[0]["label"] == "Reach"
    assert rows[0]["value"] == 1000.0
    assert rows[0]["period_comparison"] == 5.0
    assert rows[1]["period_comparison"] in (None, "")


def test_exporter_writes_chart_points(tiny_report):
    wb = load_workbook(export(tiny_report))
    rows = _read_tabular(wb[s.SHEET_CHARTS], s.CHARTS_HEADERS)
    assert len(rows) == 2
    assert rows[0]["section_nombre"] == "section_1"
    assert rows[0]["chart_type"] == "bar"
    assert rows[0]["point_label"] == "Ene"
    assert rows[1]["point_label"] == "Feb"


def test_exporter_writes_text_and_textimage_one_row_per_widget(tiny_report):
    wb = load_workbook(export(tiny_report))
    text_rows = _read_tabular(wb[s.SHEET_TEXTS], s.TEXTS_HEADERS)
    ti_rows = _read_tabular(wb[s.SHEET_TEXTIMAGES], s.TEXTIMAGES_HEADERS)
    assert len(text_rows) == 1
    assert len(ti_rows) == 1
    assert text_rows[0]["section_nombre"] == "section_1"
    assert text_rows[0]["body"] == "Body text"
    assert ti_rows[0]["section_nombre"] == "section_1"
    assert ti_rows[0]["image_position"] == "left"


def test_exporter_empty_sheets_when_no_widgets_of_that_type(db):
    """Report sin ImageWidget → hoja Images queda vacía (solo headers)."""
    stage = make_stage()
    report = Report.objects.create(
        stage=stage,
        kind=Report.Kind.MENSUAL,
        period_start=date(2026, 4, 1),
        period_end=date(2026, 4, 30),
        title="Sin images",
        status=Report.Status.DRAFT,
    )
    section = Section.objects.create(report=report, order=1, title="Solo")
    TextWidget.objects.create(section=section, order=1, title="")

    wb = load_workbook(export(report))
    ws = wb[s.SHEET_IMAGES]
    rows = _read_tabular(ws, s.IMAGES_HEADERS)
    assert rows == []


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _read_tabular(ws, headers: list[str]) -> list[dict]:
    out = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value in (None, ""):
            continue
        row = {h: ws.cell(row=r, column=i).value for i, h in enumerate(headers, start=1)}
        out.append(row)
    return out
