"""Parser xlsx → ParsedReport + builder → Report (post sections-widgets-redesign).

Incluye un roundtrip pesado (writer → exporter → parser → builder)
que garantiza que las 4 piezas (template, export, parse, build) se
mantienen consistentes ante cualquier cambio de schema.
"""
from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from apps.reports.importers import schema as s
from apps.reports.importers.builder import build_report
from apps.reports.importers.excel_exporter import export
from apps.reports.importers.excel_parser import parse
from apps.reports.importers.excel_writer import build_template, to_bytes
from apps.reports.models import (
    ChartDataPoint,
    ChartWidget,
    ImageWidget,
    KpiGridWidget,
    KpiTile,
    Report,
    Section,
    TableRow,
    TableWidget,
    TextImageWidget,
    TextWidget,
    TopContentItem,
    TopContentsWidget,
    TopCreatorItem,
    TopCreatorsWidget,
)
from apps.reports.tests.factories import make_stage


# ---------------------------------------------------------------------------
# Basic parse tests (no DB)
# ---------------------------------------------------------------------------
def test_parse_empty_template_returns_errors_for_missing_required():
    parsed, errors = parse(build_template().getvalue())
    assert parsed is None
    # Faltan tipo, fecha_inicio, fecha_fin
    required_keys = {e.column for e in errors if e.sheet == s.SHEET_REPORTE}
    assert "tipo" in required_keys
    assert "fecha_inicio" in required_keys
    assert "fecha_fin" in required_keys


def test_parse_missing_sheet_fatal():
    wb = load_workbook(build_template())
    wb.remove(wb["KpiGrids"])
    buf = BytesIO()
    wb.save(buf)
    parsed, errors = parse(buf.getvalue())
    assert parsed is None
    assert any(e.sheet == "KpiGrids" and e.reason == "hoja faltante" for e in errors)


def test_parse_corrupt_xlsx():
    parsed, errors = parse(b"not-an-xlsx")
    assert parsed is None
    assert errors[0].sheet == "(workbook)"


def test_parse_invalid_date():
    xlsx = _fill_minimal_xlsx(fecha_inicio="no-es-fecha")
    parsed, errors = parse(xlsx)
    assert parsed is None
    assert any(e.column == "fecha_inicio" and "inválida" in e.reason for e in errors)


def test_parse_invalid_kind_enum():
    xlsx = _fill_minimal_xlsx(tipo="Trimestral")  # no está en enum
    parsed, errors = parse(xlsx)
    assert parsed is None
    assert any(e.column == "tipo" and "'Trimestral' inválido" in e.reason for e in errors)


def test_parse_section_nombre_duplicated():
    """Mismo nombre de section dos veces → error."""
    wb = load_workbook(build_template())
    _fill_scalars(wb[s.SHEET_REPORTE])
    ws_sec = wb[s.SHEET_SECTIONS]
    ws_sec.cell(row=2, column=1, value="seccion")
    ws_sec.cell(row=2, column=3, value="stack")
    ws_sec.cell(row=2, column=4, value=1)
    ws_sec.cell(row=3, column=1, value="seccion")
    ws_sec.cell(row=3, column=3, value="stack")
    ws_sec.cell(row=3, column=4, value=2)

    buf = BytesIO()
    wb.save(buf)
    parsed, errors = parse(buf.getvalue())
    assert parsed is None
    assert any("duplicado" in e.reason.lower() for e in errors)


def test_parse_widget_references_unknown_section():
    """Un widget apunta a section_nombre inexistente → error."""
    wb = load_workbook(build_template())
    _fill_scalars(wb[s.SHEET_REPORTE])
    _add_section(wb[s.SHEET_SECTIONS], nombre="real", order=1)

    ws_texts = wb[s.SHEET_TEXTS]
    ws_texts.cell(row=2, column=1, value="fantasma")   # section_nombre
    ws_texts.cell(row=2, column=2, value=1)             # widget_orden

    buf = BytesIO()
    wb.save(buf)
    parsed, errors = parse(buf.getvalue())
    assert parsed is None
    assert any("fantasma" in e.reason for e in errors)


def test_parse_widget_orden_duplicated_within_section():
    """Dos widgets con el mismo (section_nombre, widget_orden) → error."""
    wb = load_workbook(build_template())
    _fill_scalars(wb[s.SHEET_REPORTE])
    _add_section(wb[s.SHEET_SECTIONS], nombre="sec1", order=1)

    ws_texts = wb[s.SHEET_TEXTS]
    ws_texts.cell(row=2, column=1, value="sec1")
    ws_texts.cell(row=2, column=2, value=1)
    ws_texts.cell(row=2, column=4, value="body 1")

    # Second widget in same sheet, same section + orden → duplicate
    ws_texts.cell(row=3, column=1, value="sec1")
    ws_texts.cell(row=3, column=2, value=1)
    ws_texts.cell(row=3, column=4, value="body 2")

    buf = BytesIO()
    wb.save(buf)
    parsed, errors = parse(buf.getvalue())
    assert parsed is None
    assert any("widget_orden" in e.column and "duplicado" in e.reason for e in errors)


def test_parse_image_ref_missing_in_bundle():
    wb = load_workbook(build_template())
    _fill_scalars(wb[s.SHEET_REPORTE])
    _add_section(wb[s.SHEET_SECTIONS], nombre="sec1", order=1)
    ws_img = wb[s.SHEET_IMAGES]
    ws_img.cell(row=2, column=1, value="sec1")
    ws_img.cell(row=2, column=2, value=1)
    ws_img.cell(row=2, column=4, value="hero.jpg")  # imagen obligatoria
    buf = BytesIO()
    wb.save(buf)

    parsed, errors = parse(buf.getvalue(), available_images=frozenset())
    assert parsed is None
    assert any("no presente en images/" in e.reason for e in errors)


def test_parse_kpigrid_denormalized_consistency_violation():
    wb = load_workbook(build_template())
    _fill_scalars(wb[s.SHEET_REPORTE])
    _add_section(wb[s.SHEET_SECTIONS], nombre="sec1", order=1)
    ws = wb[s.SHEET_KPIGRIDS]
    # Two rows with same (section, widget_orden) but different widget_title.
    ws.cell(row=2, column=1, value="sec1")   # section_nombre
    ws.cell(row=2, column=2, value=1)         # widget_orden
    ws.cell(row=2, column=3, value="KPIs A")  # widget_title
    ws.cell(row=2, column=4, value=1)         # tile_orden
    ws.cell(row=2, column=5, value="Reach")   # label
    ws.cell(row=2, column=6, value=1000)      # value
    ws.cell(row=3, column=1, value="sec1")
    ws.cell(row=3, column=2, value=1)
    ws.cell(row=3, column=3, value="KPIs B")  # different title
    ws.cell(row=3, column=4, value=2)
    ws.cell(row=3, column=5, value="Eng")
    ws.cell(row=3, column=6, value=5.5)
    buf = BytesIO()
    wb.save(buf)
    parsed, errors = parse(buf.getvalue())
    assert parsed is None
    assert any("difiere del usado antes" in e.reason for e in errors)


# ---------------------------------------------------------------------------
# Roundtrip: writer → exporter → parser → builder
# ---------------------------------------------------------------------------
@pytest.fixture
def full_report(db):
    """Report con 1 Section y 1 widget de cada tipo (8 total), sin imágenes."""
    stage = make_stage()
    report = Report.objects.create(
        stage=stage,
        kind=Report.Kind.MENSUAL,
        period_start=date(2026, 4, 1),
        period_end=date(2026, 4, 30),
        title="Full",
        intro_text="Intro",
        conclusions_text="Concl",
        status=Report.Status.DRAFT,
    )
    section = Section.objects.create(report=report, order=1, title="Main", layout="stack")

    TextWidget.objects.create(section=section, order=1, title="Text1", body="Body")
    TextImageWidget.objects.create(
        section=section, order=2, title="TI", body="Body",
        image_position="left", columns=1,
    )

    # ImageWidget requiere imagen — le adjuntamos un contenido en memoria
    from django.core.files.base import ContentFile
    img_widget = ImageWidget(section=section, order=3, title="Hero", caption="Caption")
    img_widget.image.save("hero.jpg", ContentFile(b"fake-image"), save=False)
    img_widget.save()

    kpi = KpiGridWidget.objects.create(section=section, order=4, title="KPIs")
    KpiTile.objects.create(
        widget=kpi, order=1, label="Reach",
        value=Decimal("1000"), period_comparison=Decimal("5.0"),
    )

    tbl = TableWidget.objects.create(section=section, order=5, title="Mes")
    TableRow.objects.create(widget=tbl, order=1, is_header=True, cells=["Métrica", "Valor"])
    TableRow.objects.create(widget=tbl, order=2, cells=["reach", "500"])

    tc = TopContentsWidget.objects.create(
        section=section, order=6, title="Top posts",
        network="INSTAGRAM", period_label="abril",
    )
    TopContentItem.objects.create(widget=tc, order=1, caption="P1", source_type="ORGANIC", views=100)

    tcr = TopCreatorsWidget.objects.create(
        section=section, order=7, title="Creators",
        network="INSTAGRAM", period_label="abril",
    )
    TopCreatorItem.objects.create(widget=tcr, order=1, handle="@test", views=200)

    chart = ChartWidget.objects.create(
        section=section, order=8, title="Followers",
        network="INSTAGRAM", chart_type="bar",
    )
    ChartDataPoint.objects.create(widget=chart, order=1, label="Ene", value=Decimal("100"))
    ChartDataPoint.objects.create(widget=chart, order=2, label="Feb", value=Decimal("120"))

    return report


def test_roundtrip_export_parse_build_reconstructs_same_shape(full_report, db):
    """writer→exporter→parser→builder produce un Report funcionalmente equivalente."""
    img_filename = _image_filename(ImageWidget.objects.get(section__report=full_report))

    # 1. Export
    xlsx_bytes = export(full_report).getvalue()

    # 2. Parse
    parsed, errors = parse(xlsx_bytes, available_images={img_filename})
    assert errors == [], f"Parser encontró errores: {[e.to_dict() for e in errors]}"
    assert parsed is not None

    # 3. Build
    fake_images = {img_filename: b"fake-image-bytes"}
    new_report = build_report(parsed, fake_images, stage_id=full_report.stage.pk)

    # 4. Verify shape
    assert new_report.pk != full_report.pk
    assert new_report.kind == full_report.kind
    assert new_report.period_start == full_report.period_start
    assert new_report.period_end == full_report.period_end
    assert new_report.title == full_report.title
    assert new_report.intro_text == full_report.intro_text
    assert new_report.conclusions_text == full_report.conclusions_text
    assert new_report.status == Report.Status.DRAFT

    # Same number of sections
    assert new_report.sections.count() == full_report.sections.count()

    # One widget of each type
    new_section = new_report.sections.first()
    assert TextWidget.objects.filter(section__report=new_report).count() == 1
    assert TextImageWidget.objects.filter(section__report=new_report).count() == 1
    assert ImageWidget.objects.filter(section__report=new_report).count() == 1
    assert KpiGridWidget.objects.filter(section__report=new_report).count() == 1
    assert TableWidget.objects.filter(section__report=new_report).count() == 1
    assert TopContentsWidget.objects.filter(section__report=new_report).count() == 1
    assert TopCreatorsWidget.objects.filter(section__report=new_report).count() == 1
    assert ChartWidget.objects.filter(section__report=new_report).count() == 1

    # Nested items
    new_kpi = KpiGridWidget.objects.get(section__report=new_report)
    assert new_kpi.tiles.count() == 1
    new_tbl = TableWidget.objects.get(section__report=new_report)
    assert new_tbl.rows.count() == 2
    new_chart = ChartWidget.objects.get(section__report=new_report)
    assert new_chart.data_points.count() == 2

    # Image persisted
    new_img = ImageWidget.objects.get(section__report=new_report)
    assert new_img.image.name


def test_roundtrip_preserves_widget_order(full_report, db):
    img_filename = _image_filename(ImageWidget.objects.get(section__report=full_report))
    xlsx_bytes = export(full_report).getvalue()
    parsed, errors = parse(xlsx_bytes, available_images={img_filename})
    assert errors == []
    new_report = build_report(parsed, {img_filename: b"data"}, stage_id=full_report.stage.pk)
    section = new_report.sections.first()
    orders = list(
        section.widgets.all().order_by("order").values_list("order", flat=True)
    )
    assert orders == [1, 2, 3, 4, 5, 6, 7, 8]


def test_builder_rolls_back_on_failure(db, full_report, monkeypatch):
    """Si el builder falla a mitad de transacción, ningún Report nuevo queda en DB."""
    img_filename = _image_filename(ImageWidget.objects.get(section__report=full_report))
    xlsx_bytes = export(full_report).getvalue()
    parsed, errors = parse(xlsx_bytes, available_images={img_filename})
    assert errors == []
    pre_count = Report.objects.count()

    from apps.reports.importers import builder as b

    def boom(*args, **kwargs):
        raise RuntimeError("explotó a mitad del build")

    monkeypatch.setitem(b._BUILDERS, "ChartWidget", boom)

    with pytest.raises(RuntimeError, match="explotó"):
        build_report(parsed, {img_filename: b"x"}, stage_id=full_report.stage.pk)

    assert Report.objects.count() == pre_count


# ---------------------------------------------------------------------------
# Roundtrip: Section + Widget (from task spec)
# ---------------------------------------------------------------------------
@pytest.mark.django_db
def test_section_widget_roundtrip_export_then_parse():
    """Export a report with sections + widgets, parse the bytes back, verify structure survives."""
    from apps.reports.importers.excel_exporter import export
    from apps.reports.importers.excel_parser import parse
    from apps.reports.tests.factories import make_report

    report = make_report()
    s_obj = Section.objects.create(report=report, order=1, title="Análisis", layout=Section.Layout.STACK)
    TextWidget.objects.create(section=s_obj, order=1, title="", body="Marzo arrancó con...")
    tw = TableWidget.objects.create(section=s_obj, order=2, title="IG", show_total=True)
    TableRow.objects.create(widget=tw, order=1, is_header=True, cells=["Métrica", "Valor", "Δ"])
    TableRow.objects.create(widget=tw, order=2, cells=["ORGANIC · reach", "312000", "+9.9%"])

    xlsx_bytes = export(report).getvalue()
    parsed, errors = parse(xlsx_bytes)
    assert errors == []
    assert parsed is not None

    secs = parsed.sections
    assert len(secs) == 1
    assert secs[0].title == "Análisis"
    assert secs[0].layout == "stack"

    widgets = parsed.widgets_by_section[secs[0].nombre]
    assert len(widgets) == 2
    assert widgets[0].type_name == "TextWidget"
    assert widgets[0].fields["body"] == "Marzo arrancó con..."
    assert widgets[1].type_name == "TableWidget"
    assert widgets[1].widget_title == "IG"
    assert widgets[1].fields["widget_show_total"] is True
    assert len(widgets[1].items) == 2


def _image_filename(img_widget) -> str:
    return img_widget.image.name.rsplit("/", 1)[-1]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _fill_minimal_xlsx(
    tipo: str = "Mensual",
    fecha_inicio: str = "01/04/2026",
    fecha_fin: str = "30/04/2026",
) -> bytes:
    """Generates a template with minimal valid scalars. No sections needed for parse errors."""
    wb = load_workbook(build_template())
    ws = wb[s.SHEET_REPORTE]
    # KV rows: tipo=2, fecha_inicio=3, fecha_fin=4.
    ws.cell(row=2, column=2, value=tipo)
    ws.cell(row=3, column=2, value=fecha_inicio)
    ws.cell(row=4, column=2, value=fecha_fin)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _fill_scalars(ws) -> None:
    ws.cell(row=2, column=2, value="Mensual")
    ws.cell(row=3, column=2, value="01/04/2026")
    ws.cell(row=4, column=2, value="30/04/2026")


def _add_section(ws, nombre: str, order: int, layout: str = "stack") -> None:
    """Write a section row to the Sections sheet."""
    # Find first blank row after header
    row = 2
    while ws.cell(row=row, column=1).value is not None:
        row += 1
    ws.cell(row=row, column=1, value=nombre)
    ws.cell(row=row, column=3, value=layout)
    ws.cell(row=row, column=4, value=order)
