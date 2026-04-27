"""Parser del xlsx a `ParsedReport` (post sections-widgets-redesign).

Pure function: no toca Django ni modelos. Recibe los bytes del xlsx + el
set de filenames disponibles en el ZIP, valida el contenido y devuelve
(ParsedReport | None, List[ImporterError]). Si hay errores, retorna
ParsedReport=None — el caller no debe intentar commiteer.

Estrategia: acumular errores en vez de fail-fast.
"""
from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Callable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from . import schema as s
from .errors import ImporterError
from .parsed import ParsedReport, ParsedSection, ParsedWidget


_NOMBRE_RE = re.compile(s.NOMBRE_PATTERN)


def parse(
    xlsx_bytes: bytes,
    available_images: set[str] | frozenset[str] = frozenset(),
) -> tuple[ParsedReport | None, list[ImporterError]]:
    """Parsea el xlsx. Si `errors != []`, `ParsedReport` es None."""
    errors: list[ImporterError] = []

    try:
        wb = load_workbook(BytesIO(xlsx_bytes), data_only=True)
    except Exception:  # noqa: BLE001
        return None, [ImporterError(
            sheet="(workbook)", row=None, column=None,
            reason="xlsx corrupto o ilegible",
        )]

    # Estructural: todas las hojas deben existir.
    missing = [n for n in s.SHEETS_IN_ORDER if n not in wb.sheetnames]
    if missing:
        for name in missing:
            errors.append(ImporterError(
                sheet=name, row=None, column=None, reason="hoja faltante",
            ))
        return None, errors

    report_scalars, errors_reporte = _parse_reporte(wb[s.SHEET_REPORTE])
    errors.extend(errors_reporte)

    sections, errors_sections = _parse_sections(wb[s.SHEET_SECTIONS])
    errors.extend(errors_sections)

    # Build a set of valid section nombres for cross-reference
    section_nombres = {ps.nombre for ps in sections}

    # Parse all widget sheets
    widgets_by_section: dict[str, list[ParsedWidget]] = {}
    for sheet_name, parser_fn in _WIDGET_PARSERS.items():
        parsed_widgets, sheet_errors = parser_fn(wb[sheet_name])
        errors.extend(sheet_errors)
        for pw in parsed_widgets:
            # Validate section_nombre exists
            if pw.section_nombre not in section_nombres:
                errors.append(ImporterError(
                    sheet=sheet_name, row=None, column="section_nombre",
                    reason=(
                        f"section_nombre '{pw.section_nombre}' no existe en la "
                        "hoja Sections."
                    ),
                ))
                continue
            # Validate widget_orden uniqueness within section
            existing = widgets_by_section.setdefault(pw.section_nombre, [])
            duplicate = any(
                w.widget_orden == pw.widget_orden for w in existing
            )
            if duplicate:
                errors.append(ImporterError(
                    sheet=sheet_name, row=None, column="widget_orden",
                    reason=(
                        f"widget_orden {pw.widget_orden} duplicado en section "
                        f"'{pw.section_nombre}'. Cada widget debe tener un orden único."
                    ),
                ))
                continue
            existing.append(pw)

    # Collect image refs and validate against available_images
    image_refs: set[str] = set()
    for widgets in widgets_by_section.values():
        for pw in widgets:
            for filename in _collect_image_refs(pw):
                image_refs.add(filename)
    for filename in sorted(image_refs):
        if filename not in available_images:
            errors.append(ImporterError(
                sheet="(images)", row=None, column="imagen",
                reason=(
                    f"imagen '{filename}' referenciada en el Excel pero "
                    "no presente en images/ del ZIP."
                ),
            ))

    if errors:
        return None, errors

    assert report_scalars is not None
    return ParsedReport(
        stage_id=None,
        sections=sections,
        widgets_by_section=widgets_by_section,
        image_refs=image_refs,
        **report_scalars,
    ), errors


# ---------------------------------------------------------------------------
# Reporte (KV only — no Layout table)
# ---------------------------------------------------------------------------
def _parse_reporte(ws: Worksheet) -> tuple[dict | None, list[ImporterError]]:
    errors: list[ImporterError] = []

    raw_kv: dict[str, object] = {}
    for row_idx in range(1, ws.max_row + 1):
        key_cell = ws.cell(row=row_idx, column=1).value
        if not isinstance(key_cell, str):
            continue
        key = key_cell.strip()
        if key.startswith("#") or key == "":
            continue
        normalized = key.rstrip("*").strip()
        raw_kv[normalized] = ws.cell(row=row_idx, column=2).value

    scalars: dict[str, object] = {}
    required_by_key = {key: req for key, _t, req, _ex in s.REPORTE_KV_ROWS}
    type_by_key = {key: t for key, t, _r, _ex in s.REPORTE_KV_ROWS}

    for key, type_hint in type_by_key.items():
        value = raw_kv.get(key)
        required = required_by_key[key]

        if _is_blank(value):
            if required:
                errors.append(ImporterError(
                    sheet=s.SHEET_REPORTE, row=None, column=key,
                    reason="obligatorio",
                ))
            if type_hint == "text":
                scalars[_kv_dataclass_name(key)] = ""
            continue

        if type_hint == "enum":
            label = str(value).strip()
            if label not in s.KIND_FROM_LABEL:
                errors.append(ImporterError(
                    sheet=s.SHEET_REPORTE, row=None, column=key,
                    reason=(
                        f"valor '{label}' inválido. Esperado: "
                        + ", ".join(s.KIND_FROM_LABEL)
                    ),
                ))
                continue
            scalars[_kv_dataclass_name(key)] = s.KIND_FROM_LABEL[label]
        elif type_hint == "date":
            d = _parse_date(value)
            if d is None:
                errors.append(ImporterError(
                    sheet=s.SHEET_REPORTE, row=None, column=key,
                    reason=(
                        f"fecha inválida: '{value}'. Formatos aceptados: "
                        "DD/MM/YYYY, DD-MM-YYYY, YYYY-MM-DD."
                    ),
                ))
                continue
            scalars[_kv_dataclass_name(key)] = d
        elif type_hint == "text":
            scalars[_kv_dataclass_name(key)] = str(value)

    return (scalars if scalars else None), errors


# ---------------------------------------------------------------------------
# Sections sheet
# ---------------------------------------------------------------------------
def _parse_sections(ws: Worksheet) -> tuple[list[ParsedSection], list[ImporterError]]:
    sections: list[ParsedSection] = []
    errors: list[ImporterError] = []
    seen_nombres: set[str] = set()
    seen_orders: set[int] = set()

    for row_idx, row in _iter_data_rows(ws, s.SECTIONS_HEADERS):
        nombre_raw = str(row.get("nombre") or "").strip()
        if not nombre_raw:
            errors.append(ImporterError(
                sheet=s.SHEET_SECTIONS, row=row_idx, column="nombre",
                reason="obligatorio",
            ))
            continue
        if not _NOMBRE_RE.match(nombre_raw):
            errors.append(ImporterError(
                sheet=s.SHEET_SECTIONS, row=row_idx, column="nombre",
                reason=(
                    f"'{nombre_raw}' no cumple el patrón {s.NOMBRE_PATTERN} "
                    f"(max {s.NOMBRE_MAX_LEN} chars, a-z 0-9 _ -)."
                ),
            ))
            continue
        if nombre_raw in seen_nombres:
            errors.append(ImporterError(
                sheet=s.SHEET_SECTIONS, row=row_idx, column="nombre",
                reason=f"'{nombre_raw}' duplicado en Sections.",
            ))
            continue
        seen_nombres.add(nombre_raw)

        layout_raw = str(row.get("layout") or "stack").strip()
        if layout_raw not in s.LAYOUT_VALUES:
            errors.append(ImporterError(
                sheet=s.SHEET_SECTIONS, row=row_idx, column="layout",
                reason=(
                    f"valor '{layout_raw}' inválido. Esperado: "
                    + ", ".join(s.LAYOUT_VALUES)
                ),
            ))
            layout_raw = "stack"

        order_raw = row.get("order")
        order = _coerce_int(order_raw)
        if order is None or order < 1:
            errors.append(ImporterError(
                sheet=s.SHEET_SECTIONS, row=row_idx, column="order",
                reason=f"entero ≥ 1 esperado, recibí '{order_raw}'",
            ))
            continue
        if order in seen_orders:
            errors.append(ImporterError(
                sheet=s.SHEET_SECTIONS, row=row_idx, column="order",
                reason=f"order {order} duplicado en Sections.",
            ))
            continue
        seen_orders.add(order)

        sections.append(ParsedSection(
            nombre=nombre_raw,
            title=_str(row.get("title")),
            layout=layout_raw,
            order=order,
            instructions=_str(row.get("instructions")),
        ))

    sections.sort(key=lambda ps: ps.order)
    return sections, errors


# ---------------------------------------------------------------------------
# Widget sheet parsers
# ---------------------------------------------------------------------------
def _parse_texts(ws: Worksheet) -> tuple[list[ParsedWidget], list[ImporterError]]:
    widgets: list[ParsedWidget] = []
    errors: list[ImporterError] = []
    for row_idx, row in _iter_data_rows(ws, s.TEXTS_HEADERS):
        section_nombre, widget_orden, widget_title, ok = _parse_widget_key(
            row, s.SHEET_TEXTS, row_idx, errors,
        )
        if not ok:
            continue
        widgets.append(ParsedWidget(
            type_name="TextWidget",
            section_nombre=section_nombre,
            widget_orden=widget_orden,
            widget_title=widget_title,
            fields={"body": _str(row.get("body"))},
        ))
    return widgets, errors


def _parse_images(ws: Worksheet) -> tuple[list[ParsedWidget], list[ImporterError]]:
    widgets: list[ParsedWidget] = []
    errors: list[ImporterError] = []
    for row_idx, row in _iter_data_rows(ws, s.IMAGES_HEADERS):
        section_nombre, widget_orden, widget_title, ok = _parse_widget_key(
            row, s.SHEET_IMAGES, row_idx, errors,
        )
        if not ok:
            continue
        imagen = _str(row.get("imagen"))
        if not imagen:
            errors.append(ImporterError(
                sheet=s.SHEET_IMAGES, row=row_idx, column="imagen",
                reason="obligatorio (ImageWidget requiere imagen)",
            ))
            continue
        widgets.append(ParsedWidget(
            type_name="ImageWidget",
            section_nombre=section_nombre,
            widget_orden=widget_orden,
            widget_title=widget_title,
            fields={
                "imagen": imagen,
                "image_alt": _str(row.get("image_alt")),
                "caption": _str(row.get("caption")),
            },
        ))
    return widgets, errors


def _parse_textimages(ws: Worksheet) -> tuple[list[ParsedWidget], list[ImporterError]]:
    widgets: list[ParsedWidget] = []
    errors: list[ImporterError] = []
    for row_idx, row in _iter_data_rows(ws, s.TEXTIMAGES_HEADERS):
        section_nombre, widget_orden, widget_title, ok = _parse_widget_key(
            row, s.SHEET_TEXTIMAGES, row_idx, errors,
        )
        if not ok:
            continue
        columns = _coerce_int(row.get("columns"))
        if columns not in (None, 1, 2, 3):
            errors.append(ImporterError(
                sheet=s.SHEET_TEXTIMAGES, row=row_idx, column="columns",
                reason=f"valor '{row.get('columns')}' inválido. Esperado: 1, 2 o 3.",
            ))
        image_position = str(row.get("image_position") or "top").strip()
        if image_position not in s.IMAGE_POSITION_VALUES:
            errors.append(ImporterError(
                sheet=s.SHEET_TEXTIMAGES, row=row_idx, column="image_position",
                reason=(
                    f"valor '{image_position}' inválido. Esperado: "
                    + ", ".join(s.IMAGE_POSITION_VALUES)
                ),
            ))
            image_position = "top"
        widgets.append(ParsedWidget(
            type_name="TextImageWidget",
            section_nombre=section_nombre,
            widget_orden=widget_orden,
            widget_title=widget_title,
            fields={
                "body": _str(row.get("body")),
                "imagen": _str(row.get("imagen")),
                "image_alt": _str(row.get("image_alt")),
                "image_position": image_position,
                "columns": columns or 1,
            },
        ))
    return widgets, errors


def _parse_kpigrids(ws: Worksheet) -> tuple[list[ParsedWidget], list[ImporterError]]:
    return _parse_grouped_widget(
        ws, s.SHEET_KPIGRIDS, s.KPIGRIDS_HEADERS,
        type_name="KpiGridWidget",
        widget_field_cols=(),
        item_field_cols=(
            "tile_orden", "label", "value", "unit",
            "period_comparison", "period_comparison_label",
        ),
        numeric_item_cols={"value", "period_comparison"},
        required_item_cols=("label", "value"),
    )


def _parse_tables(ws: Worksheet) -> tuple[list[ParsedWidget], list[ImporterError]]:
    """Parser de la hoja Tables: agrupa filas por (section_nombre, widget_orden)."""
    errors: list[ImporterError] = []
    # key: (section_nombre, widget_orden) → {widget_title, show_total, rows}
    groups: dict[tuple[str, int], dict] = {}
    group_order: list[tuple[str, int]] = []

    for row_idx, row in _iter_data_rows(ws, s.TABLES_HEADERS):
        section_nombre, widget_orden, widget_title, ok = _parse_widget_key(
            row, s.SHEET_TABLES, row_idx, errors,
        )
        if not ok:
            continue

        widget_show_total = _coerce_bool(row.get("widget_show_total"))
        is_header = _coerce_bool(row.get("is_header"))
        row_orden = _coerce_int(row.get("row_orden"))
        if row_orden is None or row_orden < 1:
            errors.append(ImporterError(
                sheet=s.SHEET_TABLES, row=row_idx, column="row_orden",
                reason=f"entero ≥ 1 esperado, recibí '{row.get('row_orden')}'",
            ))
            continue

        cells_raw = [row.get(col) for col in s.TABLE_CELL_COLS]
        last_non_blank = -1
        for i, v in enumerate(cells_raw):
            if not _is_blank(v):
                last_non_blank = i
        cells = [_str(v) for v in cells_raw[: last_non_blank + 1]]

        key = (section_nombre, widget_orden)
        if key not in groups:
            groups[key] = {
                "widget_title": widget_title,
                "show_total": widget_show_total,
                "rows": [],
            }
            group_order.append(key)
        else:
            existing = groups[key]
            if existing["widget_title"] != widget_title:
                errors.append(ImporterError(
                    sheet=s.SHEET_TABLES, row=row_idx, column="widget_title",
                    reason=(
                        f"valor '{widget_title}' difiere del usado antes para "
                        f"({section_nombre}, {widget_orden}) ('{existing['widget_title']}'). "
                        "Los widget_* fields deben ser idénticos en todas las filas."
                    ),
                ))
            if existing["show_total"] != widget_show_total:
                errors.append(ImporterError(
                    sheet=s.SHEET_TABLES, row=row_idx, column="widget_show_total",
                    reason=(
                        f"valor '{widget_show_total}' difiere del usado antes para "
                        f"({section_nombre}, {widget_orden}) ('{existing['show_total']}')."
                    ),
                ))
        groups[key]["rows"].append({
            "row_orden": row_orden,
            "is_header": is_header,
            "cells": cells,
        })

    result = [
        ParsedWidget(
            type_name="TableWidget",
            section_nombre=key[0],
            widget_orden=key[1],
            widget_title=data["widget_title"],
            fields={
                "widget_show_total": data["show_total"],
            },
            items=sorted(data["rows"], key=lambda r: r["row_orden"]),
        )
        for key, data in ((k, groups[k]) for k in group_order)
    ]
    return result, errors


def _parse_charts(ws: Worksheet) -> tuple[list[ParsedWidget], list[ImporterError]]:
    widgets, errors = _parse_grouped_widget(
        ws, s.SHEET_CHARTS, s.CHARTS_HEADERS,
        type_name="ChartWidget",
        widget_field_cols=("widget_network", "chart_type"),
        item_field_cols=("point_orden", "point_label", "point_value"),
        numeric_item_cols={"point_value"},
        enum_widget_cols={
            "widget_network": (s.NETWORK_FROM_LABEL, True),
        },
        required_item_cols=("point_label", "point_value"),
    )
    # Validate chart_type
    for pw in widgets:
        ct = pw.fields.get("chart_type")
        if ct not in s.CHART_TYPE_VALUES:
            errors.append(ImporterError(
                sheet=s.SHEET_CHARTS, row=None, column="chart_type",
                reason=(
                    f"widget ({pw.section_nombre}, {pw.widget_orden}): "
                    f"chart_type '{ct}' inválido. "
                    f"Esperado: {', '.join(s.CHART_TYPE_VALUES)}."
                ),
            ))
    return widgets, errors


def _parse_topcontents(ws: Worksheet) -> tuple[list[ParsedWidget], list[ImporterError]]:
    return _parse_grouped_widget(
        ws, s.SHEET_TOPCONTENTS, s.TOPCONTENTS_HEADERS,
        type_name="TopContentsWidget",
        widget_field_cols=("widget_network", "widget_period_label"),
        item_field_cols=(
            "item_orden", "imagen", "caption", "post_url", "source_type",
            "views", "likes", "comments", "shares", "saves",
        ),
        numeric_item_cols={"views", "likes", "comments", "shares", "saves"},
        enum_widget_cols={
            "widget_network": (s.NETWORK_FROM_LABEL, True),
        },
        enum_item_cols={
            "source_type": (s.SOURCE_TYPE_FROM_LABEL, True),
        },
        required_item_cols=(),
    )


def _parse_topcreators(ws: Worksheet) -> tuple[list[ParsedWidget], list[ImporterError]]:
    return _parse_grouped_widget(
        ws, s.SHEET_TOPCREATORS, s.TOPCREATORS_HEADERS,
        type_name="TopCreatorsWidget",
        widget_field_cols=("widget_network", "widget_period_label"),
        item_field_cols=(
            "item_orden", "imagen", "handle", "post_url",
            "views", "likes", "comments", "shares",
        ),
        numeric_item_cols={"views", "likes", "comments", "shares"},
        enum_widget_cols={
            "widget_network": (s.NETWORK_FROM_LABEL, True),
        },
        required_item_cols=("handle",),
    )


_WIDGET_PARSERS: dict[str, Callable[[Worksheet], tuple[list[ParsedWidget], list[ImporterError]]]] = {
    s.SHEET_TEXTS: _parse_texts,
    s.SHEET_IMAGES: _parse_images,
    s.SHEET_TEXTIMAGES: _parse_textimages,
    s.SHEET_KPIGRIDS: _parse_kpigrids,
    s.SHEET_TABLES: _parse_tables,
    s.SHEET_CHARTS: _parse_charts,
    s.SHEET_TOPCONTENTS: _parse_topcontents,
    s.SHEET_TOPCREATORS: _parse_topcreators,
}


# ---------------------------------------------------------------------------
# Generic grouped widget parser
# ---------------------------------------------------------------------------
def _parse_grouped_widget(
    ws: Worksheet,
    sheet_name: str,
    headers: list[str],
    *,
    type_name: str,
    widget_field_cols: tuple[str, ...],
    item_field_cols: tuple[str, ...],
    numeric_item_cols: set[str],
    enum_widget_cols: dict[str, tuple[dict, bool]] | None = None,
    enum_item_cols: dict[str, tuple[dict, bool]] | None = None,
    required_item_cols: tuple[str, ...] = (),
) -> tuple[list[ParsedWidget], list[ImporterError]]:
    """Generic parser for grouped widget sheets (one row = one item)."""
    enum_widget_cols = enum_widget_cols or {}
    enum_item_cols = enum_item_cols or {}
    errors: list[ImporterError] = []
    # (section_nombre, widget_orden) → {widget_title, fields, items, first_row}
    groups: dict[tuple[str, int], dict] = {}
    group_order: list[tuple[str, int]] = []

    for row_idx, row in _iter_data_rows(ws, headers):
        section_nombre, widget_orden, widget_title, ok = _parse_widget_key(
            row, sheet_name, row_idx, errors,
        )
        if not ok:
            continue

        widget_fields = {}
        for col in widget_field_cols:
            raw = row.get(col)
            if col in enum_widget_cols:
                label_map, blank_ok = enum_widget_cols[col]
                val = _parse_enum(raw, label_map, blank_ok, sheet_name, row_idx, col, errors)
            else:
                val = _str(raw) if raw is not None else ""
            widget_fields[col] = val

        item = {}
        item_errors_this_row = 0
        for col in item_field_cols:
            raw = row.get(col)
            if col in enum_item_cols:
                label_map, blank_ok = enum_item_cols[col]
                val = _parse_enum(raw, label_map, blank_ok, sheet_name, row_idx, col, errors)
            elif col in numeric_item_cols:
                val = _parse_number(raw, sheet_name, row_idx, col, errors, required=col in required_item_cols)
                if val is None and col in required_item_cols and not _is_blank(raw):
                    item_errors_this_row += 1
            elif col in ("tile_orden", "item_orden", "point_orden"):
                val = _coerce_int(raw) if not _is_blank(raw) else None
            else:
                val = _str(raw) if raw is not None else ""
            item[col] = val

        for req_col in required_item_cols:
            if _is_blank(item.get(req_col)):
                errors.append(ImporterError(
                    sheet=sheet_name, row=row_idx, column=req_col,
                    reason="obligatorio",
                ))
                item_errors_this_row += 1

        key = (section_nombre, widget_orden)
        if key not in groups:
            groups[key] = {
                "widget_title": widget_title,
                "fields": widget_fields,
                "items": [],
                "first_row": row_idx,
            }
            group_order.append(key)
        else:
            existing = groups[key]
            if existing["widget_title"] != widget_title:
                errors.append(ImporterError(
                    sheet=sheet_name, row=row_idx, column="widget_title",
                    reason=(
                        f"valor '{widget_title}' difiere del usado antes para "
                        f"({section_nombre}, {widget_orden}) "
                        f"('{existing['widget_title']}'). "
                        "Los widget_* fields deben ser idénticos en todos los rows."
                    ),
                ))
            for col in widget_field_cols:
                existing_val = existing["fields"][col]
                new_val = widget_fields[col]
                if existing_val != new_val:
                    errors.append(ImporterError(
                        sheet=sheet_name, row=row_idx, column=col,
                        reason=(
                            f"valor '{new_val}' difiere del usado antes para "
                            f"({section_nombre}, {widget_orden}) ('{existing_val}'). "
                            "Los widget_* fields deben ser idénticos en todos los rows."
                        ),
                    ))

        if item_errors_this_row == 0:
            groups[key]["items"].append(item)

    result = [
        ParsedWidget(
            type_name=type_name,
            section_nombre=key[0],
            widget_orden=key[1],
            widget_title=data["widget_title"],
            fields=data["fields"],
            items=data["items"],
        )
        for key, data in ((k, groups[k]) for k in group_order)
    ]
    return result, errors


# ---------------------------------------------------------------------------
# Image ref collector
# ---------------------------------------------------------------------------
def _collect_image_refs(pw: ParsedWidget):
    """Yields every non-empty filename referenced by the widget."""
    img = pw.fields.get("imagen")
    if img:
        yield img
    for item in pw.items:
        if item.get("imagen"):
            yield item["imagen"]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _parse_widget_key(
    row: dict, sheet: str, row_idx: int, errors: list[ImporterError],
) -> tuple[str, int, str, bool]:
    """Parse section_nombre + widget_orden + widget_title. Returns (sn, wo, wt, ok)."""
    section_nombre = str(row.get("section_nombre") or "").strip()
    if not section_nombre:
        errors.append(ImporterError(
            sheet=sheet, row=row_idx, column="section_nombre",
            reason="obligatorio",
        ))
        return "", 0, "", False

    widget_orden_raw = row.get("widget_orden")
    widget_orden = _coerce_int(widget_orden_raw)
    if widget_orden is None or widget_orden < 1:
        errors.append(ImporterError(
            sheet=sheet, row=row_idx, column="widget_orden",
            reason=f"entero ≥ 1 esperado, recibí '{widget_orden_raw}'",
        ))
        return "", 0, "", False

    widget_title = _str(row.get("widget_title"))
    return section_nombre, widget_orden, widget_title, True


def _iter_data_rows(ws: Worksheet, headers: list[str]):
    """Yields (row_idx, {header: value}) pairs skipping fully-blank rows."""
    for row_idx in range(2, ws.max_row + 1):
        values = [ws.cell(row=row_idx, column=c).value for c in range(1, len(headers) + 1)]
        if all(_is_blank(v) for v in values):
            continue
        yield row_idx, dict(zip(headers, values))


def _parse_enum(
    raw, label_map: dict, blank_ok: bool,
    sheet: str, row_idx: int, column: str, errors: list[ImporterError],
):
    if _is_blank(raw):
        if blank_ok:
            return None
        errors.append(ImporterError(
            sheet=sheet, row=row_idx, column=column, reason="obligatorio",
        ))
        return None
    label = str(raw).strip()
    if label not in label_map:
        errors.append(ImporterError(
            sheet=sheet, row=row_idx, column=column,
            reason=(
                f"valor '{label}' inválido. Esperado: "
                + ", ".join(label_map)
            ),
        ))
        return None
    return label_map[label]


def _parse_number(
    raw, sheet: str, row_idx: int, column: str,
    errors: list[ImporterError], *, required: bool,
):
    if _is_blank(raw):
        if required:
            errors.append(ImporterError(
                sheet=sheet, row=row_idx, column=column, reason="obligatorio",
            ))
        return None
    try:
        return Decimal(str(raw))
    except (InvalidOperation, ValueError):
        errors.append(ImporterError(
            sheet=sheet, row=row_idx, column=column,
            reason=f"número esperado, recibí '{raw}'",
        ))
        return None


def _parse_date(value) -> date | None:
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if not isinstance(value, str):
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(value.strip(), fmt).date()
        except ValueError:
            continue
    return None


def _coerce_int(value) -> int | None:
    if _is_blank(value):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return None


def _coerce_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().upper() == "TRUE"
    return bool(value)


def _is_blank(v) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def _str(v) -> str:
    if v is None:
        return ""
    return str(v)


def _kv_dataclass_name(key: str) -> str:
    """Mapea la label del xlsx al field name del ParsedReport."""
    mapping = {
        "tipo": "kind",
        "fecha_inicio": "period_start",
        "fecha_fin": "period_end",
        "titulo": "title",
        "intro": "intro_text",
        "conclusiones": "conclusions_text",
    }
    return mapping[key]
