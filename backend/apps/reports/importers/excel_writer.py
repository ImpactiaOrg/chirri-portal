"""Template writer del importer xlsx (DEV-83 · Etapa 1).

`build_template()` genera el workbook vacío con las 10 hojas, headers,
dropdowns y la hoja `Instrucciones`. `build_skeleton()` genera el mismo
workbook pero sin las instrucciones pre-populadas — lo usa el exporter
cuando escribe un report existente, para no duplicar layout.
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from . import schema as s

_HEADER_FILL = PatternFill("solid", fgColor="E8EEF7")
_HEADER_FONT = Font(bold=True)
_SECTION_FONT = Font(bold=True, size=12)
_TITLE_FONT = Font(bold=True, size=14)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------
def build_template() -> BytesIO:
    """Retorna un BytesIO con el template vacío (10 hojas + dropdowns)."""
    wb = build_skeleton()
    return _to_bytes(wb)


def build_skeleton() -> Workbook:
    """Workbook con las 10 hojas, headers y dropdowns — sin filas de data.

    Expuesto para que `excel_exporter` reuse la estructura en vez de
    recrearla. El caller llena las filas y lo exporta con `_to_bytes`.
    """
    wb = Workbook()
    wb.remove(wb.active)  # sacamos la "Sheet" default

    _build_instrucciones_sheet(wb)
    _build_reporte_sheet(wb)
    for sheet_name in (
        s.SHEET_TEXTIMAGE,
        s.SHEET_IMAGENES,
        s.SHEET_KPIS,
        s.SHEET_METRICSTABLES,
        s.SHEET_TOPCONTENTS,
        s.SHEET_TOPCREATORS,
        s.SHEET_ATTRIBUTION,
        s.SHEET_CHARTS,
    ):
        _build_tabular_sheet(wb, sheet_name, s.SHEET_HEADERS[sheet_name])

    return wb


def to_bytes(wb: Workbook) -> BytesIO:
    """Serializa un Workbook a BytesIO — expuesto para el exporter."""
    return _to_bytes(wb)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------
def _build_instrucciones_sheet(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet(s.SHEET_INSTRUCCIONES)
    ws.column_dimensions["A"].width = 110

    row = 1
    row = _write_line(ws, row, "Importador de reportes · Instrucciones", font=_TITLE_FONT)
    row += 1

    for section in _INSTRUCCIONES_SECTIONS:
        row = _write_line(ws, row, section["heading"], font=_SECTION_FONT)
        for line in section["lines"]:
            row = _write_line(ws, row, line)
        row += 1  # blank line between sections

    return ws


def _build_reporte_sheet(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet(s.SHEET_REPORTE)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60

    row = 1
    ws.cell(row=row, column=1, value="# Datos del reporte").font = _SECTION_FONT
    row += 1

    # Key-value rows del Report (scalars)
    for key, _type, required, example in s.REPORTE_KV_ROWS:
        label = f"{key}{'*' if required else ''}"
        c_key = ws.cell(row=row, column=1, value=label)
        c_key.font = _HEADER_FONT
        ws.cell(row=row, column=2, value="")
        ws.cell(row=row, column=3, value=f"ej: {example}").font = Font(italic=True, color="888888")
        row += 1

    # Dropdown para `tipo` sobre B{row-6}
    tipo_row = 2  # tipo es el primer KV
    _apply_dropdown(ws, s.DROPDOWNS[(s.SHEET_REPORTE, "tipo")],
                    f"B{tipo_row}:B{tipo_row}")

    row += 1
    ws.cell(row=row, column=1, value="# Layout (orden de bloques)").font = _SECTION_FONT
    row += 1

    _write_headers_at(ws, row, s.REPORTE_LAYOUT_HEADERS)
    # No pre-llenamos filas — el writer es del template vacío.
    return ws


def _build_tabular_sheet(
    wb: Workbook, sheet_name: str, headers: list[str]
) -> Worksheet:
    ws = wb.create_sheet(sheet_name)
    _write_headers_at(ws, 1, headers)
    _auto_width(ws, headers)
    _apply_sheet_dropdowns(ws, sheet_name, headers, header_row=1, rows_below=200)
    return ws


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _write_line(ws: Worksheet, row: int, text: str, *, font: Font = None) -> int:
    cell = ws.cell(row=row, column=1, value=text)
    if font is not None:
        cell.font = font
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    return row + 1


def _write_headers_at(ws: Worksheet, row: int, headers: list[str]) -> None:
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL


def _auto_width(ws: Worksheet, headers: list[str]) -> None:
    for idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(12, len(header) + 4)


def _apply_sheet_dropdowns(
    ws: Worksheet,
    sheet_name: str,
    headers: list[str],
    *,
    header_row: int,
    rows_below: int,
) -> None:
    """Aplica DataValidation a cada columna con dropdown declarado en schema."""
    for idx, header in enumerate(headers, start=1):
        choices = s.DROPDOWNS.get((sheet_name, header))
        if not choices:
            continue
        col_letter = get_column_letter(idx)
        cell_range = f"{col_letter}{header_row + 1}:{col_letter}{header_row + rows_below}"
        _apply_dropdown(ws, choices, cell_range)


def _apply_dropdown(ws: Worksheet, choices: list[str], cell_range: str) -> None:
    # openpyxl DataValidation list formula = quoted CSV, max ~255 chars
    formula = '"' + ",".join(choices) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Valor fuera del dropdown. Usá exactamente uno de los valores permitidos."
    dv.errorTitle = "Valor inválido"
    ws.add_data_validation(dv)
    dv.add(cell_range)


def _to_bytes(wb: Workbook) -> BytesIO:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Contenido de la hoja Instrucciones
# ---------------------------------------------------------------------------
_INSTRUCCIONES_SECTIONS = [
    {
        "heading": "A. Cómo llenar el Excel",
        "lines": [
            "· Cada reporte tiene un único archivo xlsx con 10 hojas (esta + Reporte + 8 hojas de blocks).",
            "· La hoja Reporte tiene arriba la data escalar (tipo, fechas, título, intro, conclusiones) y abajo el Layout — ahí definís el orden en que aparecen los bloques.",
            "· Cada bloque tiene un 'nombre' único en todo el archivo (ej. 'intro', 'kpis_mes', 'hero'). Máximo 60 caracteres, minúsculas, números, guion y guion bajo.",
            "· El Layout referencia cada bloque por su 'nombre'. El parser deduce el tipo de la hoja donde aparece ese nombre — por eso no puede repetirse entre hojas.",
            "· Hojas con sub-items (Kpis, MetricsTables, TopContents, TopCreators, Attribution, Charts): los campos del parent (ej. block_title) se repiten en cada row del item. Agrupamos por 'nombre'.",
            "· Hojas sin items (TextImage, Imagenes): una fila = un block.",
            "· Hojas vacías significan 'no hay blocks de este tipo en el reporte'. No se rompe nada.",
            "· Fechas en formato DD/MM/YYYY. Booleans como TRUE/FALSE. Decimales con punto.",
            "· BrandFollowerSnapshot no va acá — es brand-level, se carga aparte.",
        ],
    },
    {
        "heading": "B. Cómo armar el ZIP para importar",
        "lines": [
            "· El admin acepta un archivo .zip con esta estructura:",
            "    reporte.zip",
            "    ├── reporte.xlsx           (único .xlsx en la raíz)",
            "    └── images/",
            "        ├── hero.jpg",
            "        ├── intro.jpg",
            "        └── ...",
            "· En las columnas 'imagen' (presentes en TextImage, Imagenes, TopContents, TopCreators) poné SOLO el filename (ej. hero.jpg), sin path.",
            "· Filenames case-sensitive. Extensiones aceptadas: .jpg, .jpeg, .png, .webp.",
            "· Paso a paso: 1) llená el xlsx, 2) creá la carpeta images/, 3) pegá las imágenes, 4) seleccioná ambos y 'Enviar a → Carpeta comprimida' (Windows) / 'Comprimir' (Mac), 5) subí el .zip en el admin.",
            "· Si no referenciás ninguna imagen, podés subir solo el .xlsx (sin ZIP).",
        ],
    },
    {
        "heading": "C. Regenerar el template",
        "lines": [
            "· python manage.py dump_report_template — escribe un xlsx vacío con este layout.",
            "· python manage.py dump_report_example [--report <id>] — exporta un report existente al mismo formato.",
            "· Si dejás el xlsx viejo y Django cambió el schema, bajá un template nuevo para ver los campos actuales.",
        ],
    },
    {
        "heading": "D. Para LLMs / scripts generando el ZIP desde un PDF",
        "lines": [
            "Esta sección es un contrato formal para que un LLM (Claude, ChatGPT) pueda generar un ZIP válido a partir de un PDF + este template.",
            "",
            "Constraints globales:",
            "· 'nombre' único en todo el archivo (no repetir entre hojas).",
            "· Cada 'nombre' del Layout debe existir en exactamente una hoja de blocks.",
            "· En hojas denormalizadas, los block_* fields deben ser idénticos en todos los rows con el mismo 'nombre'.",
            "· 'imagen' filenames case-sensitive, extensiones .jpg|.jpeg|.png|.webp, deben existir en images/ del ZIP.",
            "· Enums: 'tipo' ∈ {Influencer, General, Quincenal, Mensual, Cierre de etapa}. 'block_network' ∈ {Instagram, TikTok, X, (vacío)}. 'source_type' ∈ {Orgánico, Influencer, Pauta, (vacío)}. 'image_position' ∈ {left, right, top}. 'chart_type' ∈ {bar, line}. 'block_show_total' ∈ {TRUE, FALSE}.",
            "",
            "Ejemplo canónico (few-shot):",
            "· python manage.py dump_report_example (sin args) escupe un xlsx con el reporte Abril del seed — cubre los 8 block types. Es la fuente de verdad ante dudas de formato.",
            "",
            "Validación pre-submit (Etapa 2):",
            "· python manage.py validate_import <zip> chequea el ZIP sin tocar DB. Devuelve errores (hoja, fila, columna, razón). Útil como feedback loop para el LLM.",
            "",
            "Extracción de imágenes:",
            "· No se espera que el LLM extraiga imágenes del PDF. Para campos 'imagen' opcionales, dejarlos vacíos si no están disponibles; Julián las sube post-import.",
            "· Para 'Imagenes' (hoja donde 'imagen' es obligatoria), el LLM debe listar qué filenames referenció para que el usuario provea las imágenes.",
        ],
    },
]
